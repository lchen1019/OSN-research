import openpyxl


def write_line_excel(work_book, name, data):
    sheet = work_book.create_sheet(name)
    for i, row in enumerate(data):
        for j, item in enumerate(row):
            sheet.cell(i + 1, j + 1, item)


def get_matrix():
    file = open('YeastL.net', 'r', encoding='utf-8')
    data = file.readlines()
    edge = data[2374:]
    n = 2362
    for ind, row in enumerate(edge):
        edge[ind] = [row[:row[:5].index(' ')], row[5:-1]]
    matrix = [[0] * n for i in range(n)]
    for pair in edge:
        matrix[int(pair[0]) - 1][int(pair[1]) - 1] = matrix[int(pair[1]) - 1][int(pair[0]) - 1] = 1
    return matrix


def main():
    file = open('YeastL.net', 'r', encoding='utf-8')
    data = file.readlines()
    node = data[12:2373]
    edge = data[2374:]
    work_book = openpyxl.Workbook()
    # 写入node信息
    for ind, row in enumerate(node):
        index = row[:4].rfind(' ')
        d = [row[index + 1:4], row[6: -2]]
        node[ind] = d
    node.insert(0, ['Id', 'Label'])
    # write_line_excel(work_book, 'node', node)
    # 写入edge信息
    for ind, row in enumerate(edge):
        edge[ind] = [row[:row[:5].index(' ')], row[5:-1]]
    edge.insert(0, ['Source', 'Target'])
    write_line_excel(work_book, 'edge', edge)
    # 保存到excel
    work_book.save('graph.xlsx')
    get_matrix(len(node), edge)


# if __name__ == "__main__":
#     main()
