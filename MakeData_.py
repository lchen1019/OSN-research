import openpyxl


def write_line_excel(work_book, name, data):
    sheet = work_book.create_sheet(name)
    for i, row in enumerate(data):
        for j, item in enumerate(row):
            sheet.cell(i + 1, j + 1, item)


def get_matrix():
    n = 4039
    matrix = [[0] * n for i in range(n)]
    file = open('facebook_combined.txt', 'r', encoding='utf-8')
    data = file.readlines()
    edge = data[:]
    for ind, item in enumerate(edge):
        edge[ind] = item[:-1].split(' ')
    for i in edge:
        matrix[int(i[0]) - 1][int(i[1]) - 1] = matrix[int(i[1]) - 1][int(i[0]) - 1] = 1
    return matrix


def main():
    file = open('facebook_combined.txt', 'r', encoding='utf-8')
    data = file.readlines()
    node = [[(i + 1)] for i in range(4039)]
    edge = data[:]
    work_book = openpyxl.Workbook()
    # 写入node信息
    node.insert(0, ['Id', 'Label'])
    write_line_excel(work_book, 'node', node)
    # 写入edge信息
    for ind, item in enumerate(edge):
        edge[ind] = item[:-1].split(' ')
    # print(i[:-1].split(' '))
    # for ind, row in enumerate(edge):
    #     edge[ind] = [row[0], row[1]]
    edge.insert(0, ['Source', 'Target'])
    write_line_excel(work_book, 'edge', edge)
    # # 保存到excel
    work_book.save('graph.xlsx')


if __name__ == "__main__":
    main()